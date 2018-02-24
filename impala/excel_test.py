#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
# @Time    : 2018/2/23 14:36
# @Author  : Frank
# @Site    : 
# @File    : excel_test.py
# @Software: PyCharm
"""
import numpy as np
import pandas as pd

# # prepare for data
# data = np.arange(1,34).reshape((11,3))
# data_df = pd.DataFrame(data)
#
# # change the index and column name
# data_df.columns = [u'2月9日',u'2月9日（123269）',u'2月9日（124770）']
# data_df.index = [u'历史首次充值人数',u'历史首次充值用户充值金额(当天充值)',u'竞猜题目数',u'房间登录用户数',u'房间登录用户充值人数',u'房间登录用户充值金额（去除红星闪闪）',
#                  u'房间登录用户首次充值人数',u'房间登录用户首次充值金额',u'房间登录用户数新增注册人数',u'房间登录用户数新增注册充值人数',u'房间登录用户数新增注册充值金额']
#
# # create and writer pd.DataFrame to excel
# writer = pd.ExcelWriter('Save_Excel.xlsx')
# data_df.to_excel(writer,u'充值注册',float_format='%.5f') # float_format 控制精度
# writer.save()


#使用pandas读取excel文件
from openpyxl import load_workbook

xls_file=pd.ExcelFile('Save_Excel.xlsx')
xls_file.sheet_names#显示出读入excel文件中的表名字
table1=xls_file.parse(u'充值注册')
# table2=xls_file.parse('second_sheet')
print table1
print type(table1)
# xlsx_file=pd.ExcelFile("./demo.xlsx")
# x1=xlsx_file.parse(0)
# x2=xlsx_file.parse(1)

# excel文件的写出
# data.to_excel("abc.xlsx",sheet_name="abc",index=False,header=True)  #该条语句会运行失败，原因在于写入的对象是np数组而不是DataFrame对象,只有DataFrame对象才能使用to_excel方法。

# pd.DataFrame(data).to_excel("abc.xlsx",sheet_name="123",index=False,header=True)

# excel文件和pandas的交互读写，主要使用到pandas中的两个函数,一个是pd.ExcelFile函数,一个是to_excel函数


pos=pd.read_excel('Save_Excel.xlsx')
print pos
pos=pd.read_excel('Save_Excel.xlsx',header=None)
print pos

# -------------
"""xlwt、wlrd只能读写xls文件，而不能操作xlsx文件
   Python使用openpyxl读写excel文件,这是一个第三方库，可以处理xlsx格式的Excel文件。pip install openpyxl安装。如果使用Aanconda，应该自带了
"""
# https://www.jianshu.com/p/892023680381?utm_campaign=maleskine&utm_content=note&utm_medium=seo_notes&utm_source=recommendation

"""2007版以前的Excel（xls结尾的），需要使用xlrd读，xlwt写。
   2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写。
"""

import xlrd
import xlwt
import openpyxl
def write03Excel(path):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet(u"2003测试表")
    value = [[u"名称", u"价格", u"出版社", u"语言"],
             [u"如何高效读懂一本书", "22.3", u"机械工业出版社", u"中文"],
             [u"暗时间", "32.4", u"人民邮电出版社", u"中文"],
             [u"拆掉思维里的墙", "26.7", u"机械工业出版社", u"中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入数据成功！")


def read03Excel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t")
        print()


def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = u'2007测试表'

    value = [[u"名称", u"价格", u"出版社", u"语言"],
             [u"如何高效读懂一本书", "22.3", u"机械工业出版社", u"中文"],
             [u"暗时间", "32.4", u"人民邮电出版社", u"中文"],
             [u"拆掉思维里的墙", "26.7", u"机械工业出版社", u"中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name(u'2007测试表')

    for row in sheet.rows:
        for cell in row:
            print (cell.value, "\t")
        print()


file_2003 = '2003.xls'
file_2007 = '2007.xlsx'

# write03Excel(file_2003)
# read03Excel(file_2003)

data = pd.read_excel('Save_Excel.xlsx', sheetname=u'充值注册')
# col_data = list(data.ix[:, 5])  # 获取除表头外开始的第五列数据
row_data = list(data.ix[5,:])  # 获取除表头外开始的第五行数据
writer = pd.ExcelWriter( 'Save_Excel.xlsx', engine='openpyxl')
book = load_workbook('Save_Excel.xlsx')
writer.book = book
result = pd.DataFrame(row_data)
result.to_excel(writer,sheet_name='a', index=False)
writer.save()

# import sys
# reload(sys)
# print sys.getdefaultencoding()
# sys.setdefaultencoding('utf-8')
# print sys.getdefaultencoding()
"""直接写excel报错
   'ascii' codec can't encode characters in position 0-1: ordinal not in range(128)
   用getdefaultencoding()函数查询Python的编码为'ascii'编码
   查了资料发现Python默认的编码方式为'ascii'编码而不是'utf-8'编码,将PyCharm的编码格式改为'utf-8'也没有用
   后来找到了解决方法：在manage.py文件的开头加上如下代码：
    import sys
    reload(sys)
    sys.setdefaultencoding('utf-8')
关于为什么要reload(sys)？
   因为这里的import语句可能不是sys的第一次导入语句，可能是第2、3次进行sys模块的import，这里是一个对sys的引用，只能reload才能进行重新加载。
为什么要重新加载，而直接引用过来则不能调用该函数？
   因为setdefaultencoding()函数在被系统调用后被删除了（所以如果不重新加载，在Pycharm中显示setdefaultencoding()函数不存在），所以通过import引用进来时其实已经没有了，
   所以必须reload一次sys模块，这样setdefaultencoding()才会为可用，才能在代码里修改解释器当前的字符编码。
"""
# write07Excel(file_2007)
# read07Excel(file_2007)