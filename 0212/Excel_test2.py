#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
# @Time    : 2018/2/26 14:45
# @Author  : Frank
# @Site    : 
# @File    : Excel_test2.py
# @Software: PyCharm
"""
"""xlwt、wlrd只能读写xls文件，而不能操作xlsx文件
   Python使用openpyxl读写excel文件,这是一个第三方库，可以处理xlsx格式的Excel文件,不支持xls格式。pip install openpyxl安装。如果使用Aanconda，应该自带了
"""
# https://www.jianshu.com/p/892023680381?utm_campaign=maleskine&utm_content=note&utm_medium=seo_notes&utm_source=recommendation

"""2007版以前的Excel（xls结尾的），需要使用xlrd读，xlwt写。
   2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写。
"""

import xlrd
import xlwt
import openpyxl
def write03Excel(path):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet(u"2003测试表")
    value = [[u"名称", u"价格", u"出版社", u"语言"],
             [u"如何高效读懂一本书", "22.3", u"机械工业出版社", u"中文"],
             [u"暗时间", "32.4", u"人民邮电出版社", u"中文"],
             [u"拆掉思维里的墙", "26.7", u"机械工业出版社", u"中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入数据成功！")


def read03Excel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t")
        print()


def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = u'2007测试表'

    value = [[u"名称", u"价格", u"出版社", u"语言"],
             [u"如何高效读懂一本书", "22.3", u"机械工业出版社", u"中文"],
             [u"暗时间", "32.4", u"人民邮电出版社", u"中文"],
             [u"拆掉思维里的墙", "26.7", u"机械工业出版社", u"中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name(u'2007测试表')

    for row in sheet.rows:
        for cell in row:
            print (cell.value, "\t")
        print()


file_2003 = '2003.xls'
file_2007 = '2007.xlsx'

# write03Excel(file_2003)
# read03Excel(file_2003)

import sys
reload(sys)
print sys.getdefaultencoding()
sys.setdefaultencoding('utf-8')
print sys.getdefaultencoding()
"""直接写excel报错
   'ascii' codec can't encode characters in position 0-1: ordinal not in range(128)
   用getdefaultencoding()函数查询Python的编码为'ascii'编码
   查了资料发现Python默认的编码方式为'ascii'编码而不是'utf-8'编码,将PyCharm的编码格式改为'utf-8'也没有用
   后来找到了解决方法：在.py文件的开头加上如下代码：
    import sys
    reload(sys)
    sys.setdefaultencoding('utf-8')
关于为什么要reload(sys)？
   因为这里的import语句可能不是sys的第一次导入语句，可能是第2、3次进行sys模块的import，这里是一个对sys的引用，只能reload才能进行重新加载。
为什么要重新加载，而直接引用过来则不能调用该函数？
   因为setdefaultencoding()函数在被系统调用后被删除了（所以如果不重新加载，在Pycharm中显示setdefaultencoding()函数不存在），所以通过import引用进来时其实已经没有了，
   所以必须reload一次sys模块，这样setdefaultencoding()才会为可用，才能在代码里修改解释器当前的字符编码。
"""
write07Excel(file_2007)
read07Excel(file_2007)